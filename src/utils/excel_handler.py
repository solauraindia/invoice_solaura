import os
from openpyxl import load_workbook
from datetime import datetime
from .number_to_words import convert_to_words

class ExcelInvoiceGenerator:
    def __init__(self, template_path):
        self.template_path = template_path
        self.workbook = None
        self.worksheet = None

    def load_template(self):
        self.workbook = load_workbook(self.template_path)
        self.worksheet = self.workbook.active

    def write_value(self, cell_ref, value):
        """Write value directly to worksheet"""
        try:
            self.worksheet[cell_ref].value = value
        except:
            # If direct assignment fails, try writing to the first cell of the merged range
            cell = self.worksheet[cell_ref]
            if cell.parent.merged_cells:
                for merged_range in cell.parent.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        min_col, min_row, max_col, max_row = merged_range.bounds
                        self.worksheet.cell(row=min_row, column=min_col).value = value
                        break

    def generate_invoice(self, data, calculations):
        self.load_template()

        try:
            # Company details
            self.write_value('C4', data['company_name'])
            self.write_value('C5', f"PAN: {data['pan']}")
            self.write_value('C6', f"GST: {data['gst']}")

            # Address
            address_parts = data['address'].split(',')
            if len(address_parts) > 2:
                self.write_value('K13', ', '.join(address_parts[:2]) + ',')
                self.write_value('K14', ', '.join(address_parts[2:4]) + ',')
                self.write_value('K15', ', '.join(address_parts[4:]))
            else:
                self.write_value('K13', data['address'])

            # Date and period
            current_date = datetime.now().strftime("%d-%m-%Y")
            self.write_value('K9', f"Date of Invoice: {current_date}")
            
            # Format period dates with selected year
            from_date = datetime.strptime(f"{data['period_from']} {data['year']}", "%B %Y").strftime("01-%m-%Y")
            to_date = datetime.strptime(f"{data['period_to']} {data['year']}", "%B %Y").strftime("31-%m-%Y")
            self.write_value('H20', f"{from_date} to {to_date}")

            # Project details - handle multiple projects with proper formatting
            projects = data['project'].split(' and ')
            if len(projects) > 1:
                # If multiple projects, write first project in C20 and rest in C21
                self.write_value('C20', projects[0] + ',')
                self.write_value('C21', ' and '.join(projects[1:]))
            else:
                # If single project, write in C20 only
                self.write_value('C20', data['project'])
                self.write_value('C21', None)

            # Calculations
            total_issued = calculations['total_issued']
            net_rate = calculations['net_rate']
            calc_value = float((total_issued * net_rate))
            calc_value_with_rate = float((calc_value * 0.09))
            total_invoice_value = float((calc_value + 2 * calc_value_with_rate))

            # Invoice details
            self.write_value('C31', f"Sale of renewable attributes for I-REC ({total_issued:.4f} units at INR {net_rate:.4f} per unit)")
            self.write_value('G31', f"{calc_value:.4f}")
            self.write_value('I31', f"{calc_value_with_rate:.4f}")
            self.write_value('K31', f"{calc_value_with_rate:.4f}")
            
            # Totals
            self.write_value('G37', f"{calc_value:.4f}")
            self.write_value('I37', f"{calc_value_with_rate:.4f}")
            self.write_value('K37', f"{calc_value_with_rate:.4f}")
            self.write_value('O38', f"{total_invoice_value:.4f}")

            # Amount in words
            self.write_value('G40', convert_to_words(total_invoice_value))

        except Exception as e:
            print(f"Error in generate_invoice: {str(e)}")
            raise

    def save(self, output_path):
        try:
            self.workbook.save(output_path)
        except Exception as e:
            print(f"Error saving workbook: {str(e)}")
            raise 